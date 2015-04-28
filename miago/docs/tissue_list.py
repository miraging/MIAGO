# get the tissue list

from openpyxl import load_workbook

wb = load_workbook(filename = r'data\transfering_assumption.xlsx')
ws = wb.get_sheet_by_name(name = 'test')
tissue_list = [] # list of tissue


updateFile = open('tissue_list','a')


for i in range(2,1021):
    tissue =  ws.cell('J'+ str(i)).value
    tissue = str(tissue).strip()
    if tissue is not None:
        if tissue not in tissue_list:
            tissue_list.append(str(tissue))
            updateFile.write(tissue+'\n')
    print tissue+' '+str(i)