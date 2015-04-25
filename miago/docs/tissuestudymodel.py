# this file is to get the OWL code for tissue and study model in instance
# input: the sheet:assay_tissue_organism_mapping in table transfering_assumption.xlsx
# output: noconclusion  tissuestudymodel_individual.owl

"""
NamedIndividual pattern:

conclusion "conclusion based on data measured by assay"(MIAGO_0000104) assay
assay "hasEvaluant"(MIAGO_0000031) tissue
tissue "is par of"(BFO_0000051) study model

"""

import conclusion_assay_list
import math

from openpyxl import load_workbook

wb = load_workbook(filename = r'data\transfering_assumption.xlsx')
ws = wb.get_sheet_by_name(name = 'assay_tissue_organism_mapping')

ins_study_model = []
ins_tissue = []
ins_conclusion = []
line_list = []

updateFile = open('noclonclusion','a')
updateFile1 = open('tissuestudymodel_individual.owl','a')

for i in range(2,1006):
    print i
    study_model = ws.cell('A'+ str(i)).value
    tissue = ws.cell('B'+ str(i)).value
    conclusion_up = ws.cell('C'+ str(i)).value
    conclusion_down = ws.cell('D'+ str(i)).value

    if study_model is not None:
        study_model = str(study_model).strip()
    else:
        updateFile.write('no record in '+str(i)+'\n')
        continue
    if tissue is not None:
        tissue = str(tissue).strip()
    else:
        tissue = 'no tissue'
    
    
    if conclusion_up is not None:
        conclusion_up = str(conclusion_up).strip()
        ins_conclusion.append(str(conclusion_up))
        ins_study_model.append(str(study_model))
        ins_tissue.append(str(tissue))
        line_list.append(str(i))
    else:
        if conclusion_down is not None:
            conclusion_down = str(conclusion_down).strip()
            ins_conclusion.append(str(conclusion_down))
            ins_study_model.append(str(study_model))
            ins_tissue.append(str(tissue))
            line_list.append(str(i))
        else: 
            updateFile.write('"'+study_model+'";"'+tissue+'";"'+str(i)+'"\n')
 
organism = ['human','mouse','rat','pig','Drosophila']
organism_ID = ['NCBITaxon_9606','NCBITaxon_10090','NCBITaxon_10116','NCBITaxon_9823','NCBITaxon_7227']

updateFile.write('\n\n=========\ins_conclusion not matching, check the input table\n==========\n')

varID = 2930

for j in range(len(ins_conclusion)) :
    conclusion = ins_conclusion[j]
    if conclusion in conclusion_assay_list.conclusion_text_list:
        p = conclusion_assay_list.conclusion_text_list.index(conclusion)
        assay_ID = conclusion_assay_list.assay_ID_list[p]
        tissue = ins_tissue[j]+ ' ' + line_list[j]
        line_number = line_list[j]
        #line = line_list[j]  error: list index out of range, but the files are correct
        study_model = ins_study_model[j]+ ' ' + line_list[j]
        
        # determine the mother term for study model
        for k in range(len(organism)) :
            if organism[k] in study_model :
                mother_study_model = organism_ID[k]
            else:
                mother_study_model = 'OBI_0100026'

        #TODO: determine the mother term for tissue
        
		
        # write the OWL for study_model -- only create the term
        varID = varID + 1
        len1 = int(math.log10(varID)) + 1
        string_val = "0" * (7-len1)
        study_model_ID = string_val + str(varID)
        updateFile1.write('    <!-- http://purl.obolibrary.org/obo/MIAGO_'+str(study_model_ID)+' -->\n')
        updateFile1.write('    <owl:NamedIndividual rdf:about="&obo;MIAGO_'+str(study_model_ID)+'">\n')
        updateFile1.write('        <rdf:type rdf:resource="&obo;'+mother_study_model+'"/>\n')
        updateFile1.write('        <rdfs:label xml:lang="en">'+study_model+'</rdfs:label>\n')
        updateFile1.write('    </owl:NamedIndividual>\n\n\n')

        # write the OWL for tissue -- add pattern: tissue "is par of"(BFO_0000051) study model
        varID = varID + 1
        len1 = int(math.log10(varID)) + 1
        string_val = "0" * (7-len1)
        tissue_ID = string_val + str(varID)    
        updateFile1.write('    <!-- http://purl.obolibrary.org/obo/MIAGO_'+str(tissue_ID)+' -->\n')
        updateFile1.write('    <owl:NamedIndividual rdf:about="&obo;MIAGO_'+str(tissue_ID)+'">\n')
        updateFile1.write('        <rdf:type rdf:resource="&obo;BFO_0000040"/>\n')
        updateFile1.write('        <rdfs:label xml:lang="en">'+tissue+'</rdfs:label>\n')
        updateFile1.write('        <obo:BFO_0000050 rdf:resource="&obo;MIAGO_'+str(study_model_ID)+'"/>\n')
        updateFile1.write('    </owl:NamedIndividual>\n\n\n')

        # add pattern: assay "hasEvaluant"(MIAGO_0000031) tissue
        updateFile1.write('    <!-- http://purl.obolibrary.org/obo/'+str(assay_ID)+' -->\n')
        updateFile1.write('    <owl:NamedIndividual rdf:about="&obo;'+str(assay_ID)+'">\n')
        updateFile1.write('        <rdf:type rdf:resource="&obo;BFO_0000040"/>\n')
        updateFile1.write('        <MIAGO_0000031 rdf:resource="&obo;MIAGO_'+str(tissue_ID)+'"/>\n')
        updateFile1.write('    </owl:NamedIndividual>\n\n\n')
        
    else:
        updateFile.write(conclusion+'\n')


updateFile.close()
updateFile1.close()
