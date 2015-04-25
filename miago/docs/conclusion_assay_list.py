# get the conclusion_assay_list

from openpyxl import load_workbook

wb = load_workbook(filename = r'data\conclusion_assay_query.xlsx')
ws = wb.get_sheet_by_name(name = 'Sheet1')
conclusion_ID_list = [] # list of conclusion ID
conclusion_text_list =[]
assay_ID_list = [] 
assay_list = []

for i in range(2,610):
    conclusion_ID = ws.cell('A'+ str(i)).value
    conclusion_text = ws.cell('B'+ str(i)).value
    assay_ID = ws.cell('C'+ str(i)).value
    assay = ws.cell('D'+ str(i)).value

    conclusion_ID_list.append(str(conclusion_ID))
    conclusion_text_list.append(str(conclusion_text))
    assay_ID_list.append(str(assay_ID))
    assay_list.append(str(assay))
    #print i