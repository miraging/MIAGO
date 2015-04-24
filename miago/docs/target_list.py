from openpyxl import load_workbook

wb = load_workbook(filename = 'data\mapping table.xlsx')
ws = wb.get_sheet_by_name( name = 'final2')
target_protein_list = []
target_gene_list = []
target_gene_name_list =[]
NCBI_gene_list = []

for i in range(2, 168):
    #print i
    target_protein_ID = ws.cell('A'+ str(i)).value
    target_gene_name = ws.cell('C'+ str(i)).value
    target_gene_ID = ws.cell('D'+ str(i)).value
    NCBI_gene = ws.cell('F'+ str(i)).value
    target_protein_ID = str(target_protein_ID)
    target_gene_ID = str(target_gene_ID)
    NCBI_gene = str(NCBI_gene)
    target_protein_list.append( str(target_protein_ID) )
    target_gene_name_list.append( str(target_gene_name) )
    target_gene_list.append( str(target_gene_ID) )
    NCBI_gene_list.append( str(NCBI_gene) )

#find the position in NCBI_gene_list and then find the corresponding miRNA_ID in miRNA_ID_list

#a = "14118"
#p = NCBI_gene_list.index(a)
#print target_gene_list[p]
#print target_protein_list[p]
#print target_gene_name_list[p]