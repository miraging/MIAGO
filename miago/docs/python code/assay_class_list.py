# get the assay class list

from openpyxl import load_workbook

wb = load_workbook(filename = r'data\transfering_assumption.xlsx')
ws = wb.get_sheet_by_name(name = 'test')
assay_class_list = [] # list of disease


updateFile = open('assay_class_list','w+')


for i in range(2,1021):
    assay =  ws.cell('M'+ str(i)).value
    assay = str(assay).strip()
    if assay <> 'None':
        assay_list = assay.split(';')
        for i in range(len(assay_list)): 
            assay_sep = assay_list[i]
            if assay_sep not in assay_class_list:
                assay_class_list.append(str(assay_sep))
                updateFile.write(assay_sep+'\n')
            print assay_sep+' '+str(i)