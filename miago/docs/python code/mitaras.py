""" This file is to generate lists for the next py file to write the OWL file.
as_assump --> assumption ID list
as_uniq --> combination of gene_id and miRNA name to corespond with the as_list
as_coexpression --> coexpression conclusion list (will be Y or N)

"""
import math


from openpyxl import load_workbook

wb = load_workbook(filename = r'data\transfering_assumption.xlsx')
ws = wb.get_sheet_by_name(name = 'test')
as_assump = [] # list of assumpID
as_uniq = [] # list of ncbi_ID miRNA_name
as_coexpression = [] # list of coexpression_ID
varID = 250


updateFile = open('nocoexpression','a') 
updateFile.write('The list of no coexpression coloumn : \n')

for i in range(2,1006):
    target_name = ws.cell('E'+ str(i)).value
    #print target_name
    target_name = str(target_name).strip()

    if target_name is not None:
        miRNA_name =  ws.cell('D'+ str(i)).value
        miRNA_name = str(miRNA_name).strip()
        if miRNA_name is not None :
            NCBI_ID = ws.cell('H'+ str(i)).value
            NCBI_ID = str(NCBI_ID)
            uniq = NCBI_ID+ '_' +str(miRNA_name)
            coexpression = ws.cell('P'+ str(i)).value
        
        #p1 = miRNA_list.miRNA_name_list.index(miRNA_name)
        #miRNA_ID = miRNA_list.miRNA_ID_list[p1]
            if uniq not in as_uniq:
                as_uniq.append( str(uniq) )
                varID = varID+1
                len1 = int(math.log10(varID))+1
                string_val = "0" * (7-len1)
                assump_ID = string_val+str(varID)
                as_assump.append(str(assump_ID))
            if coexpression == "Y":
                varID = varID+1
                len1 = int(math.log10(varID))+1
                string_val = "0" * (7-len1)
                coexpression_ID = string_val+str(varID)
                as_coexpression.append(str(coexpression_ID))
            else: 
                updateFile.write(miRNA_name + ' ' + target_name + ' ' + str(i) + '\n')
        else :
            updateFile.write('no miRNA related to target : '+ target_name + ' ' + str(i)+ '\n')
    else:
        pass

updateFile.close()



#p = as_uniq.index('1021_Let-7e')
#print as_assump[p]
#print len(as_uniq)
   

    




















