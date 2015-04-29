# get the study model list

from openpyxl import load_workbook

wb = load_workbook(filename = 'transfering_assumption.xlsx')
ws = wb.get_sheet_by_name(name = 'test')
study_model_list = [] # list of study model


updateFile = open('study_model_list','a')


for i in range(2,1021):
    study_model =  ws.cell('I'+ str(i)).value
    study_model = str(study_model).strip()
    if study_model <> 'None':
        if study_model not in study_model_list:
            study_model_list.append(str(study_model))
            updateFile.write(study_model+'\n')
    print study_model+' '+str(i)