""" The pattern for miRNA target assumption (class) is:
    MIAGO_0000076  is about miRNA target
    MIAGO_0000078  is about miRNA
    MIAGO_0000045  miRNA target assumption based on
    
   
    <!-- http://purl.obolibrary.org/obo/MIAGO_0000080 -->

    <owl:Class rdf:about="&obo;MIAGO_0000080">
        <rdfs:label xml:lang="en">miR-29a targeting Ppm1d assumption</rdfs:label>
        <rdfs:subClassOf rdf:resource="&obo;MIAGO_0000079"/>
        <rdfs:subClassOf>
            <owl:Restriction>
                <owl:onProperty rdf:resource="&obo;MIAGO_0000076"/> 
                <owl:someValuesFrom rdf:resource="&obo;MIAGO_0000088"/>
            </owl:Restriction>
        </rdfs:subClassOf>
        <rdfs:subClassOf>
            <owl:Restriction>
                <owl:onProperty rdf:resource="&obo;MIAGO_0000076"/>
                <owl:someValuesFrom rdf:resource="&obo;MGI_1858214"/>
            </owl:Restriction>
        </rdfs:subClassOf>
        <rdfs:subClassOf>
            <owl:Restriction>
                <owl:onProperty rdf:resource="&obo;MIAGO_0000045"/>
                <owl:hasValue rdf:resource="&obo;MIAGO_0000100"/>
            </owl:Restriction>
        </rdfs:subClassOf>
        <rdfs:subClassOf>
            <owl:Restriction>
                <owl:onProperty rdf:resource="&obo;MIAGO_0000045"/>
                <owl:hasValue rdf:resource="&obo;MIAGO_0000099"/>
            </owl:Restriction>
        </rdfs:subClassOf>
        <rdfs:subClassOf>
            <owl:Restriction>
                <owl:onProperty rdf:resource="&obo;MIAGO_0000045"/>
                <owl:hasValue rdf:resource="&obo;MIAGO_0000103"/>
            </owl:Restriction>
        </rdfs:subClassOf>
        <rdfs:subClassOf>
            <owl:Restriction>
                <owl:onProperty rdf:resource="&obo;MIAGO_0000078"/>
                <owl:someValuesFrom rdf:resource="&obo;MIAGO_0000204"/>
            </owl:Restriction>
        </rdfs:subClassOf>
    </owl:Class>

       <!-- http://purl.obolibrary.org/obo/MIAGO_0000099 -->

    <owl:NamedIndividual rdf:about="&obo;MIAGO_0000099">
        <rdf:type rdf:resource="&obo;MIAGO_0000038"/>
        <rdfs:label xml:lang="en">miR-29a and Ppmid mRNA binding conclusion</rdfs:label>
        <obo:BFO_0000050 rdf:resource="&obo;MIAGO_0000059"/>
        <obo:MIAGO_0000104 rdf:resource="&obo;MIAGO_0000102"/>
    </owl:NamedIndividual>
    
----Individual.file

    <!-- http://purl.obolibrary.org/obo/MIAGO_0000100 -->

    <owl:NamedIndividual rdf:about="&obo;MIAGO_0000100">
        <rdf:type rdf:resource="&obo;MIAGO_0000046"/>
        <rdfs:label xml:lang="en">general bioinformatics predicted miRNA target conclusion 1(H)</rdfs:label>
        <obo:BFO_0000050 rdf:resource="&obo;MIAGO_0000059"/>
    </owl:NamedIndividual>

	    <!-- http://purl.obolibrary.org/obo/MIAGO_0000103 -->

    <owl:NamedIndividual rdf:about="&obo;MIAGO_0000103">
        <rdf:type rdf:resource="&obo;MIAGO_0000013"/>
        <rdfs:label xml:lang="en">miR-29a and Ppm1d co-expression conclusion</rdfs:label>
    </owl:NamedIndividual>

"""
import target_list
import miRNA_list
import mitaras
import math
import re

from openpyxl import load_workbook

wb = load_workbook(filename = r'data\transfering_assumption.xlsx')
ws = wb.get_sheet_by_name(name = 'test')
as_target_desc = []
as_miRNA = []
as_miRNA_name = []
as_target_protein = []
as_target_gene = []
as_target_name = []
as_prediction = []
as_assay = []
as_pmid = []
as_downstream_conclusion = []
as_upstream_conclusion = []
as_coexpression = []
as_NCBI_miRNA_name = []
line_list = []

for i in range(2,1006):
    target_name = ws.cell('E'+ str(i)).value
    #print target_name
    target_name = str(target_name).strip()
    if target_name <> 'None':
        miRNA_name =  ws.cell('D'+ str(i)).value
        miRNA_name = miRNA_name.strip()
        as_target_desc.append(str(target_name))
        validation_status =  ws.cell('F'+ str(i)).value
        NCBI_ID = ws.cell('H'+ str(i)).value
        NCBI_ID = str(NCBI_ID)
        NCBI_miRNA_name = NCBI_ID + '_'+str(miRNA_name)
        as_NCBI_miRNA_name.append(str(NCBI_miRNA_name))

        pmid = ws.cell('B'+ str(i)).value
        as_pmid.append(str(pmid))
        prediction = ws.cell('G'+ str(i)).value
        assay = ws.cell('M'+ str(i)).value
        downstream_conclusion = ws.cell('L'+ str(i)).value
        #print downstream_conclusion
        # type of MIAGO_0000037 miRNA target experimental conclusion
        upstream_conclusion = ws.cell('K'+ str(i)).value
        #print upstream_conclusion
        # type of MIAGO_0000014 miRNA expression conclusion
        coexpression = ws.cell('P'+ str(i)).value
        #print coexpression
        # type of MIAGO_0000013

        p1 = miRNA_list.miRNA_name_list.index(miRNA_name)
        miRNA_ID = miRNA_list.miRNA_ID_list[p1]
        p2 = target_list.NCBI_gene_list.index(NCBI_ID)
        target_gene_name = target_list.target_gene_name_list[p2]
        target_gene_ID = target_list.target_gene_list[p2]
        target_protein_ID = target_list.target_protein_list[p2]
        
        as_miRNA.append( str(miRNA_ID))
        as_miRNA_name.append( str(miRNA_name))
        as_target_name.append( str(target_gene_name))
        as_target_protein.append(str(target_protein_ID))
        as_target_gene.append(str(target_gene_ID))
        as_prediction.append(str(prediction))
        as_assay.append(str(assay))
        as_downstream_conclusion.append(str(downstream_conclusion))
        as_upstream_conclusion.append(str(upstream_conclusion))
        as_coexpression.append(str(coexpression))
        line_list.append(str(i))
    print i

updateFile = open(r'miRNAtarget_assumption_class_individual.owl','w+')
varID = 10620

for j in range(len(as_miRNA)):
    varID = varID +1
    len1 = int(math.log10(varID)) + 1
    string_val = "0" * (7-len1)
    pmid_ID = string_val + str(varID)

    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MIAGO_'+str(pmid_ID)+' -->\n\n')
    updateFile.write('    <owl:NamedIndividual rdf:about="&obo;MIAGO_'+str(pmid_ID)+ '">\n        <rdf:type rdf:resource="&obo;MIAGO_0000030"/>\n')
    updateFile.write('        <rdfs:label xml:lang="en">PubMed '+ as_pmid[j]+'</rdfs:label>\n')
    updateFile.write('            <dc:source>http://www.ncbi.nlm.nih.gov/pubmed/'+as_pmid[j]+'</dc:source>\n')
    updateFile.write('    </owl:NamedIndividual>\n\n\n')
    
    assay_id_list = []
    assay = as_assay[j]
    assay = str(assay).strip()
    searchObj = re.search(';', assay)
    if searchObj:
        assay_list = assay.split(';')
        for i in range(len(assay_list)):
            varID = varID +1
            len1 = int(math.log10(varID)) + 1
            string_val = "0" * (7-len1)
            assay = assay_list[i]
            assay_id = string_val + str(varID)
            assay_id_list.append(str(assay_id))
 
            updateFile.write('    <!-- http://purl.obolibrary.org/obo/MIAGO_'+str(assay_id)+' -->\n')
            updateFile.write('    <owl:NamedIndividual rdf:about="&obo;MIAGO_'+str(assay_id)+'">\n')
            updateFile.write('        <rdfs:label xml:lang="en">'+str(assay)+' '+str(line_list[j])+'</rdfs:label>\n    </owl:NamedIndividual>\n\n\n')
    else:
        if assay <> 'None':
            varID = varID +1
            len1 = int(math.log10(varID)) + 1
            string_val = "0" * (7-len1)
            assay_id = string_val + str(varID)
         
            updateFile.write('    <!-- http://purl.obolibrary.org/obo/MIAGO_'+str(assay_id)+' -->\n')
            updateFile.write('    <owl:NamedIndividual rdf:about="&obo;MIAGO_'+str(assay_id)+'">\n')
            updateFile.write('        <rdfs:label xml:lang="en">'+assay+' '+str(line_list[j])+'</rdfs:label>\n    </owl:NamedIndividual>\n\n\n')
        else:
            pass
 
    # if there is luciferase assay, then there is binding conclusion
    searchObj1 = re.search('luciferase', assay)
    if searchObj1:
        varID = varID + 1
        len1 = int(math.log10(varID))+1
        string_val = "0" * (7-len1)
        bindingconclusion_ID = string_val+str(varID)
        updateFile.write('    <!-- http://purl.obolibrary.org/obo/MIAGO_'+ bindingconclusion_ID +' -->\n\n')
        updateFile.write('    <owl:NamedIndividual rdf:about="&obo;MIAGO_'+ bindingconclusion_ID +'">\n')
        updateFile.write('        <rdf:type rdf:resource="&obo;MIAGO_0000038"/>\n')
        updateFile.write('        <rdfs:label xml:lang="en">'+as_miRNA_name[j]+' and '+as_target_name[j]+' miRNA binding conclusion</rdfs:label>\n')
        updateFile.write('        <obo:BFO_0000050 rdf:resource="&obo;MIAGO_'+str(pmid_ID)+'"/>\n')
        updateFile.write('    </owl:NamedIndividual>\n\n\n')
    else :
        bindingconclusion_ID = ''
    
    if as_downstream_conclusion[j] <> 'None':
        varID = varID + 1
        len1 = int(math.log10(varID))+1
        string_val = "0" * (7-len1)
        downconclusion_ID = string_val+str(varID)

        downconclusion = as_downstream_conclusion[j]
        miRNA_name_l = as_miRNA_name[j].split('-')
        miRNA_number = miRNA_name_l[1]
        target = as_target_desc[j]
        list_downregulates = ['decrease', 'inhibit', 'downregulate','suppress', 'repress', 'reduce']
        target_miRNA_pair = [miRNA_number,target]
        if any(x in downconclusion for x in list_downregulates):
            if all(y in downconclusion for y in target_miRNA_pair):
                mother_term = 'MIAGO_0000022'
            else:
                mother_term = 'MIAGO_0000084'
        else:
            mother_term = 'MIAGO_0000084'
        
        updateFile.write('    <!-- http://purl.obolibrary.org/obo/MIAGO_'+ downconclusion_ID +' -->\n\n')
        updateFile.write('    <owl:NamedIndividual rdf:about="&obo;MIAGO_'+ downconclusion_ID +'">\n')
        updateFile.write('        <rdf:type rdf:resource="&obo;'+str(mother_term)+'"/>\n')
        updateFile.write('        <rdfs:label xml:lang="en">'+as_downstream_conclusion[j]+'</rdfs:label>\n')
        updateFile.write('        <obo:BFO_0000050 rdf:resource="&obo;MIAGO_'+str(pmid_ID)+'"/>\n')
        if assay_id_list :
            for q in range(len(assay_id_list)):
                updateFile.write('        <obo:MIAGO_0000104 rdf:resource="&obo;MIAGO_'+assay_id_list[q]+'"/>\n')
        else:
            updateFile.write('        <obo:MIAGO_0000104 rdf:resource="&obo;MIAGO_'+assay_id+'"/>\n')
        updateFile.write('    </owl:NamedIndividual>\n\n\n')


    if as_upstream_conclusion[j] <> 'None':
    #! need a little text mining here! if up, then put into miRNA upregulated' if down, then put into miRNA downregulated'
        varID = varID + 1
        len1 = int(math.log10(varID))+1
        string_val = "0" * (7-len1)
        upconclusion_ID = string_val+str(varID)
        updateFile.write('    <!-- http://purl.obolibrary.org/obo/MIAGO_'+ upconclusion_ID +' -->\n\n')
        updateFile.write('    <owl:NamedIndividual rdf:about="&obo;MIAGO_'+ upconclusion_ID +'">\n')
        updateFile.write('        <rdf:type rdf:resource="&obo;MIAGO_0000014"/>\n')
        updateFile.write('        <rdfs:label xml:lang="en">'+as_upstream_conclusion[j]+'</rdfs:label>\n')
        updateFile.write('        <obo:BFO_0000050 rdf:resource="&obo;MIAGO_'+str(pmid_ID)+'"/>\n')
        if assay_id_list :
            for q in range(len(assay_id_list)):
                updateFile.write('        <obo:MIAGO_0000104 rdf:resource="&obo;MIAGO_'+assay_id_list[q]+'"/>\n')
        else:
            updateFile.write('        <obo:MIAGO_0000104 rdf:resource="&obo;MIAGO_'+assay_id+'"/>\n')
        updateFile.write('    </owl:NamedIndividual>\n\n\n')

    if as_coexpression[j] == 'Y':
        uniq = as_NCBI_miRNA_name[j]
        p3 = mitaras.as_uniq.index(uniq)
        coexpression_ID = mitaras.as_coexpression[p3]
        updateFile.write('    <!-- http://purl.obolibrary.org/obo/MIAGO_'+ coexpression_ID +' -->\n\n')
        updateFile.write('    <owl:NamedIndividual rdf:about="&obo;MIAGO_'+ coexpression_ID +'">\n')
        updateFile.write('        <rdf:type rdf:resource="&obo;MIAGO_0000013"/>\n')
        updateFile.write('        <rdfs:label xml:lang="en">'+as_miRNA_name[j]+' and '+as_target_name[j]+' co-expression conclusion</rdfs:label>\n')
        updateFile.write('        <obo:BFO_0000050 rdf:resource="&obo;MIAGO_'+str(pmid_ID)+'"/>\n')
        if assay_id_list :
            for q in range(len(assay_id_list)):
                updateFile.write('        <obo:MIAGO_0000104 rdf:resource="&obo;MIAGO_'+assay_id_list[q]+'"/>\n')
        else:
            updateFile.write('        <obo:MIAGO_0000104 rdf:resource="&obo;MIAGO_'+assay_id+'"/>\n')
        updateFile.write('    </owl:NamedIndividual>\n\n\n')

    if as_prediction[j] <> 'None':
        varID = varID + 1
        len1 = int(math.log10(varID))+1
        string_val = "0" * (7-len1)
        prediction_ID = string_val+str(varID)
        updateFile.write('    <!-- http://purl.obolibrary.org/obo/MIAGO_'+ prediction_ID +' -->\n\n')
        updateFile.write('    <owl:NamedIndividual rdf:about="&obo;MIAGO_' + prediction_ID +'">\n')
        updateFile.write('        <rdf:type rdf:resource="&obo;MIAGO_0000046"/>\n')
        updateFile.write('        <rdfs:label xml:lang="en">'+as_prediction[j]+' predicted miRNA target conclusion'+str(line_list[j])+'</rdfs:label>\n')
        updateFile.write('        <obo:BFO_0000050 rdf:resource="&obo;MIAGO_'+str(pmid_ID)+'"/>\n')
        updateFile.write('    </owl:NamedIndividual>\n\n\n')



    uniq = as_NCBI_miRNA_name[j]
    p3 = mitaras.as_uniq.index(uniq)
    assump_ID = mitaras.as_assump[p3]
    #miRNA = as_miRNA[j]
    #target = as_target_name[j]

    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MIAGO_'+ assump_ID +' -->\n\n')
    updateFile.write('    <owl:Class rdf:about="&obo;MIAGO_'+ assump_ID +'">\n')
    updateFile.write('        <rdfs:label xml:lang="en">'+as_miRNA_name[j]+ ' targeting '+as_target_name[j]+' assumption</rdfs:label>\n')
    updateFile.write('        <rdfs:subClassOf rdf:resource="&obo;MIAGO_0000079"/>\n')
    updateFile.write('        <rdfs:subClassOf>\n            <owl:Restriction>\n                <owl:onProperty rdf:resource="&obo;MIAGO_0000076"/> \n')
    updateFile.write('                <owl:someValuesFrom rdf:resource="&obo;'+as_target_protein[j]+'"/>\n')
    updateFile.write('            </owl:Restriction>\n        </rdfs:subClassOf>\n        <rdfs:subClassOf>\n            <owl:Restriction>\n                <owl:onProperty rdf:resource="&obo;MIAGO_0000076"/>\n')
    updateFile.write('                <owl:someValuesFrom rdf:resource="&obo;'+as_target_gene[j]+'"/>\n')
    updateFile.write('            </owl:Restriction>\n        </rdfs:subClassOf>\n')
    updateFile.write('        <rdfs:subClassOf>\n            <owl:Restriction>\n                <owl:onProperty rdf:resource="&obo;MIAGO_0000078"/>\n')
    updateFile.write('                <owl:someValuesFrom rdf:resource="&obo;'+as_miRNA[j]+'"/>\n            </owl:Restriction>\n        </rdfs:subClassOf>\n')
    if as_downstream_conclusion[j] <> 'None':
        updateFile.write('        <rdfs:subClassOf>\n            <owl:Restriction>\n                <owl:onProperty rdf:resource="&obo;MIAGO_0000045"/>\n')
        updateFile.write('                <owl:hasValue rdf:resource="&obo;MIAGO_'+downconclusion_ID+'"/>\n            </owl:Restriction>\n        </rdfs:subClassOf>\n')
    if as_upstream_conclusion[j] <> 'None':
        updateFile.write('        <rdfs:subClassOf>\n            <owl:Restriction>\n                <owl:onProperty rdf:resource="&obo;MIAGO_0000045"/>\n')
        updateFile.write('                <owl:hasValue rdf:resource="&obo;MIAGO_'+upconclusion_ID+'"/>\n            </owl:Restriction>\n        </rdfs:subClassOf>\n')
    if as_coexpression[j] == 'Y':
        updateFile.write('        <rdfs:subClassOf>\n            <owl:Restriction>\n                <owl:onProperty rdf:resource="&obo;MIAGO_0000045"/>\n')
        updateFile.write('                <owl:hasValue rdf:resource="&obo;MIAGO_'+coexpression_ID+'"/>\n            </owl:Restriction>\n        </rdfs:subClassOf>\n')
    if as_prediction[j] <> 'None':
        updateFile.write('        <rdfs:subClassOf>\n            <owl:Restriction>\n                <owl:onProperty rdf:resource="&obo;MIAGO_0000045"/>\n')
        updateFile.write('                <owl:hasValue rdf:resource="&obo;MIAGO_'+prediction_ID+'"/>\n            </owl:Restriction>\n        </rdfs:subClassOf>\n')
    if bindingconclusion_ID :
        updateFile.write('        <rdfs:subClassOf>\n            <owl:Restriction>\n                <owl:onProperty rdf:resource="&obo;MIAGO_0000045"/>\n')
        updateFile.write('                <owl:hasValue rdf:resource="&obo;MIAGO_'+bindingconclusion_ID+'"/>\n            </owl:Restriction>\n        </rdfs:subClassOf>\n')    
    updateFile.write('    </owl:Class>\n\n\n')
  

    
    
updateFile.close()

    




















