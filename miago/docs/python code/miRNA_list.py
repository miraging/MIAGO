from openpyxl import load_workbook

wb = load_workbook(filename = r'data\miRNA_ID_list.xlsx')
ws = wb.get_sheet_by_name( name = 'Sheet1')
miRNA_ID_list = []
miRNA_name_list = []

for i in range(2, 145):
    miRNA_ID = ws.cell('A'+ str(i)).value
    miRNA_name = ws.cell('B'+ str(i)).value
    miRNA_ID = str(miRNA_ID)
    miRNA_name = str(miRNA_name)
    miRNA_ID_list.append( str(miRNA_ID) )
    miRNA_name_list.append( str(miRNA_name) )

miRNA_ID_mapping = [miRNA_ID_list, miRNA_name_list]  

#find the position in miRNA_name_list and then find the corresponding miRNA_ID in miRNA_ID_list

#a = "let-7"
#p = miRNA_name_list.index(a)
#print miRNA_ID_list[p]