""" this file is to generate MIAGO ids for miRNAs in the table 'mapping table.xlsx' sheet: miRNA list
    the form should be like, starting with 0000105

    <!-- http://purl.obolibrary.org/obo/MIAGO_0000101 -->

    <owl:Class rdf:about="&obo;MIAGO_0000101">
        <rdfs:label xml:lang="en">luciferase assay</rdfs:label>
        <rdfs:subClassOf rdf:resource="&obo;SO_0000276"/>
    </owl:Class>
    
	"""


from openpyxl import load_workbook

wb = load_workbook(filename = 'data\mapping table.xlsx')
ws = wb.get_sheet_by_name( name = 'miRNA list')

for i in range(1, 144):
    varId = i + 104
    label = ws.cell('A'+ str(i)).value

    updateFile = open('miRNA_list_uri.owl','a')
    updateFile.write('    <!-- http://purl.obolibrary.org/obo/MIAGO_0000'+ str(varId) +' -->\n\n')
    updateFile.write('    <owl:Class rdf:about="&obo;MIAGO_0000'+ str(varId) +'">\n')
    updateFile.write('        <rdfs:label xml:lang="en">' + str(label)+ '</rdfs:label>\n')
    updateFile.write('        <rdfs:subClassOf rdf:resource="&obo;SO_0000276"/>\n')
    updateFile.write('    </owl:Class>\n\n\n\n')
    print i
    updateFile.close()
    