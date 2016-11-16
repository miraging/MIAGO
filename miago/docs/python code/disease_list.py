# get the disease list

from openpyxl import load_workbook

wb = load_workbook(filename = r'data\transfering_assumption.xlsx')
ws = wb.get_sheet_by_name(name = 'test')
disease_list = [] # list of disease


updateFile = open('disease_list','a')


for i in range(2,1021):
    disease =  ws.cell('O'+ str(i)).value
    disease = str(disease).strip()
    if disease is not None:
        if disease not in disease_list:
            disease_list.append(str(disease))
            updateFile.write(disease+'\n')
    print disease+' '+str(i)

updateFile.close()