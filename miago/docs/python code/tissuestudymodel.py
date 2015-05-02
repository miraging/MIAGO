# this file is to get the OWL code for tissue and study model in instance
# input: the sheet:assay_tissue_organism_mapping in table transfering_assumption.xlsx
# output: noconclusion  tissuestudymodel_individual.owl


"""
NamedIndividual pattern: (tissue used for the origin of the processed material instance)

conclusion "conclusion based on data measured by assay"(MIAGO_0000104) assay
assay "hasEvaluant"(MIAGO_0000031) processed material
tissue "processed into" processed material instance
tissue "is par of"(BFO_0000051) study model

"""

import tissue_list_final_one_list
import conclusion_assay_list
import math

from openpyxl import load_workbook

wb = load_workbook(filename = r'data\transfering_assumption.xlsx')
ws = wb.get_sheet_by_name(name = 'assay_tissue_organism_mapping')

ins_study_model = []
ins_tissue = []
ins_conclusion = []
line_list = []

updateFile = open('noconclusion','w+')
updateFile1 = open('tissuestudymodel_individual.owl','w+')

for i in range(2,1006):
    print i
    study_model = ws.cell('A'+ str(i)).value
    tissue = ws.cell('B'+ str(i)).value
    conclusion_up = ws.cell('C'+ str(i)).value
    conclusion_down = ws.cell('D'+ str(i)).value

    if study_model is not None:
        study_model = str(study_model).strip()
    else:
        updateFile.write('no record in '+str(i)+'\n')
        continue
    if tissue is not None:
        tissue = str(tissue).strip() 
    else:
        tissue = 'no tissue'
        updateFile.write(str(i)+' no tissue info. check the paper!\n')
    
    
    if conclusion_up is not None:
        conclusion_up = str(conclusion_up).strip()
        ins_conclusion.append(str(conclusion_up))
        ins_study_model.append(str(study_model))
        ins_tissue.append(str(tissue))
        line_list.append(str(i))
    else:
        if conclusion_down is not None:
            conclusion_down = str(conclusion_down).strip()
            ins_conclusion.append(str(conclusion_down))
            ins_study_model.append(str(study_model))
            ins_tissue.append(str(tissue))
            line_list.append(str(i))
        else: 
            updateFile.write('"'+study_model+'";"'+tissue+'";"'+str(i)+'"\n')
 
organism = ['human','mouse','rat','pig','Drosophila','Caenorhabditis elegans']
organism_ID = ['NCBITaxon_9606','NCBITaxon_10090','NCBITaxon_10116','NCBITaxon_9823','NCBITaxon_7227','NCBITaxon_6239']

updateFile.write('\n\n=========\ins_conclusion not matching, check the input table\n==========\n')

varID = 12930

for j in range(len(ins_conclusion)) :  # didn't think about the one conclusion mapping to multiple assay's condition?
    conclusion = ins_conclusion[j]
    
    if conclusion in conclusion_assay_list.conclusion_text_list:
      
        # determine the mother term for study model
        mother_study_model = 'OBI_0100026'
        for k in range(len(organism)) :
            if organism[k] in study_model :
                mother_study_model = organism_ID[k]
            else:
                pass
        
        # write the OWL for study_model -- only create the term
        varID = varID + 1
        len1 = int(math.log10(varID)) + 1
        string_val = "0" * (7-len1)
        study_model_ID = string_val + str(varID)
        study_model = ins_study_model[j]+ ' ' + line_list[j]
        updateFile1.write('    <!-- http://purl.obolibrary.org/obo/MIAGO_'+str(study_model_ID)+' -->\n')
        updateFile1.write('    <owl:NamedIndividual rdf:about="&obo;MIAGO_'+str(study_model_ID)+'">\n')
        updateFile1.write('        <rdf:type rdf:resource="&obo;'+mother_study_model+'"/>\n')
        updateFile1.write('        <rdfs:label xml:lang="en">'+study_model+'</rdfs:label>\n')
        updateFile1.write('    </owl:NamedIndividual>\n\n\n')
 
        
        # write the OWL for tissue -- add pattern: tissue "is par of"(BFO_0000051) study model
        #TODO: determine the mother term for tissue
        ## there is the script for assigning tissue mother terms. Maybe import.
 
        varID = varID + 1
        len1 = int(math.log10(varID)) + 1
        string_val = "0" * (7-len1)
        tissue_ID = string_val + str(varID) 
        
        #get tissue mother term from the list in tissue_list_final_one_list
        tissue_name = str(ins_tissue[j]).strip()
        if tissue_name <> 'no tissue' :
        #split tissue
            tissue_list = tissue_name.split(';')
            for i in range(len(tissue_list)):
                varID = varID + 1
                len1 = int(math.log10(varID)) + 1
                string_val = "0" * (7-len1)
                tissue_ID = string_val + str(varID) 
 
                tissue = str(tissue_list[i]).strip()
                if tissue not in tissue_list_final_one_list.tissue_instance_name_list:
                    updateFile.write(tissue +' no in tissue_final_one_list.'+line_list[j]+'\n')
                else:
                    p = tissue_list_final_one_list.tissue_instance_name_list.index(tissue)
                    tissue_mother_ID = tissue_list_final_one_list.tissue_mother_ID_list[p]
                    tissue_mother_both = tissue_list_final_one_list.tissue_mother_both_list[p]    
        # change the naming : add 'specimen' or 'culture' depends on the mother_term !!! starting from here
                    if 'UBERON' in tissue_mother_both:
                        tissue = 'processed '+ tissue + ' specimen'
                    elif 'CL' or 'CLO' in tissue_mother_both:
                        tissue = tissue + ' culture'
                    else:
                        print tissue_ID + ' ' + tissue + ' ' + line_list[j] + 'name not changed yet. TODO!'
                        updateFile.write(tissue_ID + ' ' + tissue + ' ' + line_list[j] + 'name not changed yet. TODO!\n')
                    tissue_text = tissue + ' ' + line_list[j]
                    updateFile1.write('    <!-- http://purl.obolibrary.org/obo/MIAGO_'+str(tissue_ID)+' -->\n')
                    updateFile1.write('    <owl:NamedIndividual rdf:about="&obo;MIAGO_'+str(tissue_ID)+'">\n')
                    updateFile1.write('        <rdf:type rdf:resource="&obo;'+tissue_mother_ID+'"/>\n')
                    updateFile1.write('        <rdfs:label xml:lang="en">'+tissue_text+'</rdfs:label>\n')
                    updateFile1.write('        <obo:BFO_0000050 rdf:resource="&obo;MIAGO_'+str(study_model_ID)+'"/>\n')
                    updateFile1.write('    </owl:NamedIndividual>\n\n\n')
        
                    for l in range(len(conclusion_assay_list.conclusion_text_list)):
                        if conclusion == conclusion_assay_list.conclusion_text_list[l]:
                            assay_ID = conclusion_assay_list.assay_ID_list[l]          
             
                # add pattern: assay "hasEvaluant"(MIAGO_0000031) tissue
                # ! check the multiple tissue to one assay's situation
                            updateFile1.write('    <!-- http://purl.obolibrary.org/obo/'+str(assay_ID)+' -->\n')
                            updateFile1.write('    <owl:NamedIndividual rdf:about="&obo;'+str(assay_ID)+'">\n')
                            updateFile1.write('        <obo:MIAGO_0000031 rdf:resource="&obo;MIAGO_'+str(tissue_ID)+'"/>\n')
                            updateFile1.write('    </owl:NamedIndividual>\n\n\n')
                        else:
                            pass
        
    else:
        updateFile.write(conclusion+ ' ' + line_list[j]+'\n')  # put the conclusion into no_conclusion list file.


updateFile.close()
updateFile1.close()
